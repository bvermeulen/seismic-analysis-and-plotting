from datetime import date, timedelta
import inspect
import re
import glob
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from shapely.geometry import Point
from shapely.geometry.polygon import Polygon
from shapely.ops import cascaded_union
from geopandas import GeoSeries, GeoDataFrame, read_file, overlay
import matplotlib.pyplot as plt
from PIL import Image
import contextily as ctx

from Utils.plogger import Logger
from Utils.utils import string_to_value_or_nan


PREFIX = r'autoseis_data\OUT_'
geo_shapefile = './areas_shapes/geo_shapefile.shp'
EPSG_31256_adapted = "+proj=tmerc +lat_0=0 +lon_0=16.33333333333333"\
                     " +k=1 +x_0=+500000 +y_0=0 +ellps=bessel "\
                     "+towgs84=577.326,90.129,463.919,5.137,1.474,5.297,2.4232 "\
                     "+units=m +no_defs"
EPSG_OSM = 3857
EPSG_WGS84 = 4326

ASK_DATE = 'date (YYMMDD) [q - quit]: '

logger = Logger.getlogger()

def get_date():
    _date = input(ASK_DATE)
    if _date in ['q', 'Q']:
        exit()
    _date = date(int(_date[0:2])+2000,
                 int(_date[2:4]),
                 int(_date[4:6]))

    return _date


def daterange(start_date, end_date):
    assert start_date <= end_date, f'start date {start_date} must be less than '\
                                   f'or equal to end date {end_date}'

    for n in range(int((end_date - start_date).days)+1):
        yield start_date + timedelta(n)


def get_date_range():
    start_date = input(ASK_DATE)
    if start_date in ['q', 'Q']:
        exit()

    end_date = input(ASK_DATE)
    if end_date in ['q', 'Q']:
        exit()

    start_date = date(int(start_date[0:2])+2000,
                      int(start_date[2:4]),
                      int(start_date[4:6]))
    end_date = date(int(end_date[0:2])+2000,
                    int(end_date[2:4]),
                    int(end_date[4:6]))

    if start_date > end_date:
        print('incorrect date range')
        start_date = -1
        end_date = -1

    return start_date, end_date


def swath_selection(swaths_selected=None):
    ''' Selection of swath. Swath information taken from the file:
           Points+Lines_SW_24_stay.xlsx
        Manual inpput check if there exists at least one valid swath otherwise
        repeat the input

        retrieves the extent for each swath, transforms it to Easting,Northing and
        unions the swaths to one polynom

        Parameters: None
        Returns:
        :swaths: list of selelected swaths, empty list if no swath is selected
        :swaths_pnt_polygon: union of selected swaths polygon in points (RL, RP)
        :swaths_geo_polygon: union of selected swaths polygon in (easting, northing)

    '''
    swath_file = r'./Points+Lines_SW_24_stay.xlsx'
    swath_df = pd.read_excel(swath_file, skiprows=5)
    valid_swaths = swath_df['Swath'].tolist()
    
    swaths = []
    if swaths_selected is None:
        valid = False
        while not valid:
            _swaths = [int(num[0]) for num in re.finditer(r'\d+', input(
                'Swaths to be included: [0 for all]: '))]
            if len(_swaths) == 1 and _swaths[0] == 0:
                valid = True
                break

            for swath in _swaths:
                if swath in valid_swaths:
                    swaths.append(swath)

            if swaths:
                valid = True
    else:
        if len(swaths_selected) == 1 and swaths_selected[0] == 0:
            pass
        else:
            for swath in swaths_selected:
                if swath in valid_swaths:
                    swaths.append(swath)

    swaths_pnt_polygon = []
    swaths_geo_polygon = []
    for swath in swaths:
        sd = swath_df[swath_df['Swath'] == swath].iloc[0]

        point1 = (sd['1st RL'], sd['1st GP'])
        point2 = (sd['1st RL'], sd['last GP'])
        point3 = (sd['last RL'], sd['last GP'])
        point4 = (sd['last RL'], sd['1st GP'])
        _polygon = Polygon([point1, point2, point3, point4])
        swaths_pnt_polygon.append(_polygon)

        point1_coord = transformation(point1)
        point2_coord = transformation(point2)
        point3_coord = transformation(point3)
        point4_coord = transformation(point4)
        _polygon = Polygon([point1_coord, point2_coord, point3_coord, point4_coord])
        swaths_geo_polygon.append(_polygon)

    return swaths, cascaded_union(swaths_pnt_polygon), cascaded_union(swaths_geo_polygon)

# 26-8-2019": replaced url: http://tile.stamen.com/terrain/tileZ/tileX/tileY.png'
def add_basemap_osm(ax, plot_area, zoom,
                    url='http://tile.stamen.com/terrain/{z}/{x}/{y}.png'):
    '''  load the map in OpenStreetMap format from url

         Parameters:
         :input:
            ax: matplotlib axes
            plot_area: extent of the map to be taken
            zoom: zoom factor for map (13 seems to be good)
            url: url with map information
        :output: none
    '''
    logger.info(f'url: {url}, plot_area: {plot_area}')
    basemap, extent = ctx.bounds2img(*plot_area, zoom=zoom, url=url)
    ax.imshow(basemap, extent=extent, interpolation='bilinear')


def add_basemap_local(ax):
    '''  load the map in picture format and the georeference information from the jgW file
         in the same folder; the crs has to be the same as the data
         Parameters:
         :input: ax
         :output: none
    '''
    MAP_FILE = r'BackgroundMap/3D_31256.jpg'
    Image.MAX_IMAGE_PIXELS = 2000000000

    # read the map image file and set the extent
    fname_jgW = MAP_FILE[:-4] + '.jgW'
    basemap = plt.imread(MAP_FILE)
    cols = basemap.shape[0]
    rows = basemap.shape[1]

    with open(fname_jgW, 'tr') as jgw:
        dx = float(jgw.readline())
        _ = jgw.readline()  # to do with rotation of the map to be ignored
        _ = jgw.readline()  # to do with rotation of the map to be ignored
        dy = float(jgw.readline())
        x_min = float(jgw.readline())
        y_max = float(jgw.readline())

    x_max = x_min + rows * dx
    y_min = y_max + cols * dy
    logger.info(f'filename: {MAP_FILE}, (rows: {rows}, colums: {cols}), \n'\
                f'extent map crs:{EPSG_31256_adapted}: \n {(x_min, x_max, y_min, y_max)}')

    ax.imshow(basemap, extent=(x_min, x_max, y_min, y_max), interpolation='bilinear')

class GeoData:
    '''  method for handling Geo data '''
    def __init__(self):
        self.geo_df = None

    def read_geo_data(self, _date):
        read_is_valid = False
        _geo_file = ''.join([PREFIX,
                             f'{_date.year:04}', f'{_date.month:02}', f'{_date.day:02}',
                             f'*.xlsx'])
        _geo_file = glob.glob(_geo_file)
        logger.info(f'filename: {_geo_file}')

        if len(_geo_file) == 1:
            try:
                self.geo_df = pd.read_excel(_geo_file[0])
                self.date = _date
                self.add_bat_days_in_field_to_df()
                read_is_valid = True

            except FileNotFoundError:
                logger.info(f'{inspect.stack()[0][3]} - Exception FileNotFoundError": '\
                            f'{_geo_file[0]}')
        else:
            pass

        if read_is_valid:
            return True
        else:
            return False

    def get_geo_df(self):
        return self.geo_df

    def add_bat_days_in_field_to_df(self):
        days_in_field = []
        for _, row in self.geo_df.iterrows():
            bat_start = str(row['BATSTART'])
            try:
                _year = int(bat_start[0:4])
                _julianday = int(bat_start[4:7])
                _date_bat_start = date(_year, 1, 1) + timedelta(_julianday - 1)
            except ValueError:
                logger.info(f'{inspect.stack()[0][3]} '\
                            f'- Exception ValueError - bat_start: {bat_start}')
                _date_bat_start = date(1900, 1, 1)

            bat_start_new = str(row['BATSTART_NEW'])
            try:
                _year = int(bat_start_new[0:4])
                _julianday = int(bat_start_new[4:7])
                _date_bat_start_new = date(_year, 1, 1) + timedelta(_julianday - 1)
            except ValueError:
                _date_bat_start_new = date(1900, 1, 1)

            _date_bat_start = max(_date_bat_start_new, _date_bat_start)
            if _date_bat_start != date(1900, 1, 1):
                _days_in_field = (self.date - _date_bat_start).days
            else:
                _days_in_field = np.NaN

            days_in_field.append(_days_in_field)

        # add the columns to the dataframe
        self.geo_df['days_in_field'] = days_in_field


    def filter_geo_data_by_swaths(self, swaths_selected=None, swaths_only=False,
                                  source_boundary=False):
        ''' method to select geo_data depending on swaths selected
            Parameters:
            :self: instance of GeoData
            :swaths_only: boolean - True if swath selection is required, default False
            Returns:
            :_date: date in datetime.date format
            :swaths: list of selected swaths
            :self.geo_df: pandas dataframe of reveivers
            :swaths_pnt_polygon: union of selected swaths polygon in points (RL, RP)
            :swaths_geo_polygon: union of selected swaths polygon in (easting, northing)
        '''
        swaths, swaths_pnt_polygon, swaths_geo_polygon = swath_selection(
            swaths_selected=swaths_selected)
        bnd_gdf = read_file(geo_shapefile)
        bnd_gdf.crs = EPSG_31256_adapted
        rcv_bnd_gdf = bnd_gdf[bnd_gdf['OBJECTID'] == 1]
        src_bnd_gdf = bnd_gdf[bnd_gdf['OBJECTID'] > 1]
        swaths_bnd_gdf = GeoDataFrame(geometry=GeoSeries(swaths_geo_polygon),)
        swaths_bnd_gdf.crs = EPSG_31256_adapted
        if swaths_pnt_polygon:
            swaths_bnd_gdf = overlay(rcv_bnd_gdf, swaths_bnd_gdf, how='intersection')
        else:
            swaths_bnd_gdf = rcv_bnd_gdf

        if source_boundary and swaths != []:
            src_bnd_gdf = overlay(src_bnd_gdf, swaths_bnd_gdf, how='intersection')
            swaths_bnd_gdf = overlay(swaths_bnd_gdf, src_bnd_gdf, how='union')
        else:
            pass

        if not swaths_only and swaths_pnt_polygon:
            for index, row in self.geo_df.iterrows():
                # check if point is within swath selection
                line = string_to_value_or_nan(str(row['STATIONVIX'])[0:4], 'int')
                station = string_to_value_or_nan(str(row['STATIONVIX'])[4:8], 'int')
                point = Point(line, station)
                if swaths_pnt_polygon.contains(point) or \
                   swaths_pnt_polygon.intersects(point):
                    pass  # point is in or on the polygon
                else:
                    self.geo_df = self.geo_df.drop([index])

            self.geo_df = self.geo_df.reset_index(drop=True)

        else:
            pass

        return swaths, self.geo_df, swaths_pnt_polygon, swaths_bnd_gdf


def df_to_excel(df, filename, sheet_name='Sheet1', startrow=None,
                append=True, **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
    :filename: File path or existing ExcelWriter
               (Example: '/path/to/file.xlsx')
    :df: dataframe to save to workbook
    :sheet_name: Name of sheet which will contain DataFrame.
                 (default: 'Sheet1')
    :startrow: upper left cell row to dump data frame.
               Per default (startrow=None) calculate the last row
               in the existing DF and write to the next row...
    :append: True data appended on existing sheet; False new sheet will be created
    :to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                      [can be dictionary]

    Returns: None
    """
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')  #pylint: disable=abstract-class-instantiated

    if append:
        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

        except FileNotFoundError:
            # file does not exist yet, we will create it
            logger.info(f'{inspect.stack()[0][3]} - Exception FileNotFoundError '\
                        f'{filename}')

        if startrow is None:
            startrow = 0

    else:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    valid = False
    while not valid:
        try:
            writer.save()
            valid = True
        except PermissionError:
            if input(f'Close file {filename} and enter or quit [Q] ') in ['q', 'Q']:
                exit()


def transformation(point):
    ''' transformation from RL-RP to Easting/ Northting
        Origin point taken from SPS receiver point 3400-4213
        Azimuth angle of prospect is 30 degrees

        Parameters:
        :point: a tuple of receiver line and receiver point (RL, RP)
        Returns:
        :transformed_point: a tuple of transformed point in (Easting, Northing)
    '''
    azimuth = (np.pi * 30 / 180)  # converted to radians
    dx_crossline = 10.0 * np.cos(azimuth)
    dy_crossline = -10.0 * np.sin(azimuth)
    dx_inline = -10.0 * np.sin(azimuth)
    dy_inline = -10.0 * np.cos(azimuth)
    POINT_0 = (3400., 4213.)
    COORD_0 = (491074., 5358167.)

    transformed_point = ((point[0] - POINT_0[0]) * dx_crossline +
                         (point[1] - POINT_0[1]) * dx_inline + COORD_0[0],
                         (point[0] - POINT_0[0]) * dy_crossline +
                         (point[1] - POINT_0[1]) * dy_inline + COORD_0[1])
    return transformed_point


def offset_transformation(inline_offset, crossline_offset):
    '''  transformation from inline_offset, crossline offset to delta_easting,
         delta_northing
         Azimuth angle of prospect is 30 degrees

         Parameters:
         :inline_offset: inline offset in meters, negative along azimuth vector (float)
         :crossline_offset: crossline offset in meters, positive counterclockwise (float)
         :Returns:
         :dx: (m) change in x direction (float)
         :dy: (m) change in y direction (float)
    '''
    azimuth = (np.pi * 30 / 180)  # converted to radians
    dx_crossline = crossline_offset * np.cos(azimuth)
    dy_crossline = -crossline_offset * np.sin(azimuth)
    dx_inline = -inline_offset * np.sin(azimuth)
    dy_inline = -inline_offset * np.cos(azimuth)

    return dx_inline + dx_crossline, dy_inline + dy_crossline
