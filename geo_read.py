import glob
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import date, timedelta
from Utils.plogger import Logger


PREFIX = r'autoseis_data\OUT_'
EXCEL_SUMMARY_FILE = 'geo_summary.xlsx'
thresholdtype1 = 20
thresholdtype2 = 10


class GeoData:
    '''  method for handling Geo data '''
    def __init__(self):
        self.logger = Logger.getlogger()

    def read_geo_datafile(self, _date):
        read_is_valid = False
        _geo_file = ''.join([PREFIX, 
                            f'{_date.year:04}', f'{_date.month:02}', f'{_date.day:02}', 
                            '*.xlsx'])
        _geo_file = glob.glob(_geo_file)
        self.logger.info(f'filename: {_geo_file}')

        if len(_geo_file) != 1:
            pass
        else:
            try:
                self.geo_df = pd.read_excel(_geo_file[0])
                self.date = _date
                self.add_bat_status_to_df()
                read_is_valid = True

            except FileNotFoundError:
                pass
        
        if read_is_valid:
            return True, self.geo_df
        else:    
            return False, None

    def add_bat_status_to_df(self):
        days_in_field_type1 = []
        days_in_field_type2 = []
        days_over_threshold = []
        batteries = zip(self.geo_df['Battype'].tolist(), self.geo_df['BATSTART'].tolist())
        for battery in batteries:
            if not pd.isnull(battery[1]):
                _battery = str(battery[1])
                _year = int(_battery[0:4])
                _julianday = int(_battery[4:7])
                _date_in_field = date(_year, 1, 1) + timedelta(_julianday)
                _days_in_field = (self.date - _date_in_field).days
            else:
                _days_in_field = np.NaN

            if not pd.isnull(_days_in_field) and int(battery[0]) in [1, 2]:
                if int(battery[0]) == 1:
                    days_in_field_type1.append(_days_in_field)
                    days_in_field_type2.append(np.NaN)
                    days_over_threshold.append(_days_in_field - thresholdtype1)
                elif int(battery[0]) == 2:
                    days_in_field_type1.append(np.NaN)
                    days_in_field_type2.append(_days_in_field)
                    days_over_threshold.append(_days_in_field - thresholdtype2)
            else:
                days_in_field_type1.append(np.NaN)
                days_in_field_type2.append(np.NaN)
                days_over_threshold.append(np.NaN)

        assert len(days_in_field_type1) == len(days_in_field_type2), "length days_in_field_types not equal"
        assert len(days_over_threshold) == len(days_in_field_type2), "length threshold not equal"

        # add the columns to the dataframe
        self.geo_df['days_in_field_type1'] = days_in_field_type1
        self.geo_df['days_in_field_type2'] = days_in_field_type2
        self.geo_df['days_over_threshold'] = days_over_threshold


def get_date():
    _date = input('date (YYMMDD) [q - quit]: ')
    if _date in ['q', 'Q']:
        exit()
    _date = date(int(_date[0:2])+2000, 
                 int(_date[2:4]), 
                 int(_date[4:6]))

    return _date


def daterange(start_date, end_date):
    for n in range(int ((end_date - start_date).days)):
        yield start_date + timedelta(n)


def get_date_range():
    start_date = input('date (YYMMDD) [q - quit]: ')
    end_date = input('date (YYMMDD) [q - quit]: ')
    if start_date in ['q', 'Q'] or end_date in ['q', 'Q']:
        exit()
    start_date = date(int(start_date[0:2])+2000, 
                      int(start_date[2:4]), 
                      int(start_date[4:6]))
    end_date = date(int(end_date[0:2])+2000, 
                      int(end_date[2:4]), 
                      int(end_date[4:6]))
    end_date += timedelta(1)

    if start_date >= end_date:
        print('incorrect date range')
        start_date = -1
        end_date = -1

    return start_date, end_date


def append_df_to_excel(df, filename=EXCEL_SUMMARY_FILE, sheet_name='Sheet1', 
                       startrow=None,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
    :filename: File path or existing ExcelWriter
               (Example: '/path/to/file.xlsx')
    :df: dataframe to save to workbook
    :sheet_name: Name of sheet which will contain DataFrame.
                 (default: 'Sheet1')
    :startrow: upper left cell row to dump data frame.
               Per default (startrow=None) calculate the last row
               in the existing DF and write to the next row...
    :to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                      [can be dictionary]

    Returns: None
    """
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

            # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()